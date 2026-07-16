from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable


def flatten_trace_judge(evaluation: Any) -> dict[str, Any]:
    ev = evaluation if isinstance(evaluation, dict) else {}
    return {
        'trace_judge_correctness': ev.get('correctness'),
        'trace_judge_completeness': ev.get('completeness'),
        'trace_judge_relevance': ev.get('relevance'),
        'trace_judge_groundedness': ev.get('groundedness'),
        'trace_judge_final_score': ev.get('final_score'),
        'trace_judge_explanation': ev.get('explanation'),
        'trace_judge_root_cause': ev.get('root_cause'),
        'trace_judge_mode': ev.get('judge_mode'),
    }


def flatten_for_excel(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, set)):
        return ', '.join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def flatten_row(row: dict[str, Any]) -> dict[str, Any]:
    return {k: flatten_for_excel(v) for k, v in row.items()}


def _collect_keys(rows: Iterable[dict[str, Any]]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _populate_worksheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(['empty'])
        return
    flat_rows = [flatten_row(r) for r in rows]
    keys = _collect_keys(flat_rows)
    ws.append(keys)
    for row in flat_rows:
        ws.append([row.get(k, '') for k in keys])


def _add_sheet(wb: Any, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    _populate_worksheet(ws, rows)


def _summary_rows(summary: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key, value in summary.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                rows.append({'section': key, 'metric': sub_key, 'value': flatten_for_excel(sub_value)})
        elif isinstance(value, list):
            rows.append({'section': key, 'metric': 'items', 'value': flatten_for_excel(value)})
        else:
            rows.append({'section': 'run', 'metric': key, 'value': flatten_for_excel(value)})
    return rows


def _aggregate_rows(single_turn: list[dict[str, Any]], multi_turn: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def avg(rows_in: list[dict[str, Any]], field: str) -> float | None:
        vals = [float(r[field]) for r in rows_in if r.get(field) is not None and str(r.get(field)) != '']
        return round(sum(vals) / len(vals), 4) if vals else None

    if single_turn:
        rows.append({
            'group': 'single_turn',
            'count': len(single_turn),
            'avg_latency_sec': avg(single_turn, 'latency_sec'),
            'avg_tool_f1': avg(single_turn, 'tool_f1'),
            'avg_trace_judge_final_score': avg(single_turn, 'trace_judge_final_score'),
            'avg_batch_judge_correctness': avg(single_turn, 'batch_judge_correctness'),
            'failure_count': sum(1 for r in single_turn if r.get('failure_reasons')),
        })
    if multi_turn:
        rows.append({
            'group': 'multi_turn',
            'count': len(multi_turn),
            'avg_elapsed_ms': avg(multi_turn, 'elapsed_ms'),
            'avg_trace_judge_final_score': avg(multi_turn, 'trace_judge_final_score'),
            'pass_rate': round(
                sum(1 for r in multi_turn if r.get('gate_passed')) / len(multi_turn),
                4,
            ) if multi_turn else None,
        })
    by_category: dict[str, list[dict[str, Any]]] = {}
    for row in multi_turn:
        cat = str(row.get('category') or 'Uncategorized')
        by_category.setdefault(cat, []).append(row)
    for cat, cat_rows in sorted(by_category.items()):
        rows.append({
            'group': f'category:{cat}',
            'count': len(cat_rows),
            'pass_rate': round(
                sum(1 for r in cat_rows if r.get('gate_passed')) / len(cat_rows),
                4,
            ),
            'avg_trace_judge_final_score': avg(cat_rows, 'trace_judge_final_score'),
        })
    return rows


def write_bulk_evaluation_workbook(
    *,
    output_path: Path,
    summary: dict[str, Any],
    single_turn_rows: list[dict[str, Any]] | None = None,
    multi_turn_rows: list[dict[str, Any]] | None = None,
) -> bool:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return False

    single_turn_rows = single_turn_rows or []
    multi_turn_rows = multi_turn_rows or []

    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'Summary'
    _populate_worksheet(ws0, _summary_rows(summary))
    _add_sheet(wb, 'SingleTurn', single_turn_rows)
    _add_sheet(wb, 'MultiTurn', multi_turn_rows)
    _add_sheet(wb, 'Aggregates', _aggregate_rows(single_turn_rows, multi_turn_rows))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True


def _session_aggregate_rows(
    session_rollups: list[dict[str, Any]],
    turn_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def avg(rows_in: list[dict[str, Any]], field: str) -> float | None:
        vals = [float(r[field]) for r in rows_in if r.get(field) is not None and str(r.get(field)) != '']
        return round(sum(vals) / len(vals), 4) if vals else None

    by_env: dict[str, list[dict[str, Any]]] = {}
    for row in turn_rows:
        env = str(row.get('eval_environment') or 'unknown')
        by_env.setdefault(env, []).append(row)
    for env, env_rows in sorted(by_env.items()):
        rows.append({
            'group': f'eval_environment:{env}',
            'count': len(env_rows),
            'avg_latency_sec': avg(env_rows, 'latency_sec'),
            'avg_elapsed_ms': avg(env_rows, 'elapsed_ms'),
            'avg_trace_judge_final_score': avg(env_rows, 'trace_judge_final_score'),
            'avg_batch_judge_correctness': avg(env_rows, 'batch_judge_correctness'),
        })

    by_type: dict[str, list[dict[str, Any]]] = {}
    for row in session_rollups:
        st = str(row.get('session_type') or 'unknown')
        by_type.setdefault(st, []).append(row)
    for st, st_rows in sorted(by_type.items()):
        rows.append({
            'group': f'session_type:{st}',
            'count': len(st_rows),
            'avg_turns': avg(st_rows, 'turns_count'),
            'avg_trace_judge_final_score': avg(st_rows, 'avg_trace_judge_final_score'),
            'avg_batch_judge_correctness': avg(st_rows, 'avg_batch_judge_correctness'),
            'pass_rate': round(
                sum(1 for r in st_rows if r.get('gate_passed')) / len(st_rows),
                4,
            ) if st_rows else None,
            'failure_count': sum(int(r.get('failure_count') or 0) for r in st_rows),
        })

    by_policy: dict[str, list[dict[str, Any]]] = {}
    for row in session_rollups:
        policy = str(row.get('dynamic_policy') or '')
        if not policy:
            continue
        by_policy.setdefault(policy, []).append(row)
    for policy, policy_rows in sorted(by_policy.items()):
        rows.append({
            'group': f'dynamic_policy:{policy}',
            'count': len(policy_rows),
            'avg_trace_judge_final_score': avg(policy_rows, 'avg_trace_judge_final_score'),
            'pass_rate': round(
                sum(1 for r in policy_rows if r.get('gate_passed')) / len(policy_rows),
                4,
            ),
        })

    by_category: dict[str, list[dict[str, Any]]] = {}
    for row in turn_rows:
        cat = str(row.get('category') or 'Uncategorized')
        by_category.setdefault(cat, []).append(row)
    for cat, cat_rows in sorted(by_category.items()):
        rows.append({
            'group': f'category:{cat}',
            'count': len(cat_rows),
            'avg_trace_judge_final_score': avg(cat_rows, 'trace_judge_final_score'),
            'avg_batch_judge_correctness': avg(cat_rows, 'batch_judge_correctness'),
            'gate_pass_rate': round(
                sum(1 for r in cat_rows if r.get('gate_passed')) / len(cat_rows),
                4,
            ) if any('gate_passed' in r for r in cat_rows) else None,
        })

    by_tier: dict[str, list[dict[str, Any]]] = {}
    for row in session_rollups:
        tier = str(row.get('difficulty_tier') or 'Unknown')
        by_tier.setdefault(tier, []).append(row)
    for tier, tier_rows in sorted(by_tier.items()):
        rows.append({
            'group': f'difficulty_tier:{tier}',
            'count': len(tier_rows),
            'avg_trace_judge_final_score': avg(tier_rows, 'avg_trace_judge_final_score'),
        })
    return rows


def _analytics_sheet_rows(analytics_summary: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Build optional analytics sheets from analytics_summary.json."""
    run_m = analytics_summary.get('run_metrics') if isinstance(analytics_summary.get('run_metrics'), dict) else {}
    trace_m = analytics_summary.get('trace_metrics') if isinstance(analytics_summary.get('trace_metrics'), dict) else {}
    fail_m = analytics_summary.get('failure_analytics') if isinstance(analytics_summary.get('failure_analytics'), dict) else {}

    run_kpis = [{'metric': k, 'value': flatten_for_excel(v)} for k, v in run_m.items() if k != 'extra']
    failures = [
        {'failure_type': ft, 'count': c}
        for ft, c in (fail_m.get('by_failure_type') or {}).items()
    ]
    trace_nodes = [
        {'node': n.get('node'), 'avg_duration_sec': n.get('avg_duration_sec')}
        for n in (trace_m.get('slowest_nodes') or [])
    ]
    experiments = analytics_summary.get('experiment_comparison')
    exp_rows = []
    if isinstance(experiments, dict):
        for k, v in (experiments.get('deltas') or {}).items():
            exp_rows.append({'metric': k, 'delta': v})
    return {
        'RunKPIs': run_kpis,
        'Failures': failures,
        'TraceNodes': trace_nodes,
        'Experiments': exp_rows,
    }


def write_session_evaluation_workbook(
    *,
    output_path: Path,
    summary: dict[str, Any],
    session_rollups: list[dict[str, Any]] | None = None,
    turn_rows: list[dict[str, Any]] | None = None,
    analytics_summary: dict[str, Any] | None = None,
) -> bool:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return False

    session_rollups = session_rollups or []
    turn_rows = turn_rows or []

    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'Summary'
    _populate_worksheet(ws0, _summary_rows(summary))
    _add_sheet(wb, 'Sessions', session_rollups)
    _add_sheet(wb, 'Turns', turn_rows)
    _add_sheet(wb, 'Aggregates', _session_aggregate_rows(session_rollups, turn_rows))

    if analytics_summary:
        for sheet_name, rows in _analytics_sheet_rows(analytics_summary).items():
            if rows:
                _add_sheet(wb, sheet_name, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True


def load_jsonl_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows
